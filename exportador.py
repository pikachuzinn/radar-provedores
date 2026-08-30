"""
exportador.py — Exportação dos resultados para CSV e/ou Excel.

Depende apenas de bibliotecas padrão (csv) para o formato CSV, e do
openpyxl (via pandas) para o formato Excel. O pandas não é obrigatório
para o modo CSV — o script funciona sem ele caso apenas CSV seja usado.
"""

import csv
import logging
from datetime import datetime
from pathlib import Path

from config import COLUNAS_SAIDA, DIRETORIO_SAIDA

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Helpers internos
# ---------------------------------------------------------------------------

def _garantir_diretorio(caminho: Path) -> None:
    """Cria o diretório de saída se ainda não existir."""
    caminho.mkdir(parents=True, exist_ok=True)


def _nome_arquivo_base(prefixo: str = "provedores") -> str:
    """Gera um nome de arquivo com timestamp para evitar sobrescrever resultados."""
    agora = datetime.now().strftime("%Y%m%d_%H%M%S")
    return f"{prefixo}_{agora}"


def _preparar_linhas(dados: list[dict]) -> list[dict]:
    """
    Converte a lista de dicts internos para a ordem e os rótulos
    definidos em COLUNAS_SAIDA.
    """
    linhas = []
    for item in dados:
        linha = {
            rotulo: item.get(chave, "")
            for chave, rotulo in COLUNAS_SAIDA.items()
        }
        linhas.append(linha)
    return linhas


# ---------------------------------------------------------------------------
# Exportação para CSV
# ---------------------------------------------------------------------------

def exportar_csv(dados: list[dict], diretorio: str = DIRETORIO_SAIDA) -> Path:
    """
    Salva os resultados em um arquivo CSV com codificação UTF-8 BOM
    (compatível com Excel ao abrir direto pelo duplo clique no Windows).

    Args:
        dados: Lista de dicts retornada por BuscadorProvedores.buscar_todos().
        diretorio: Pasta de destino (criada automaticamente se não existir).

    Returns:
        Caminho absoluto do arquivo criado.

    Raises:
        PermissionError: Se não houver permissão de escrita no diretório.
        OSError: Para outros erros de E/S.
    """
    pasta = Path(diretorio)
    _garantir_diretorio(pasta)

    caminho = pasta / f"{_nome_arquivo_base()}.csv"
    cabecalhos = list(COLUNAS_SAIDA.values())
    linhas = _preparar_linhas(dados)

    try:
        # utf-8-sig = UTF-8 com BOM — evita caracteres especiais trocados no Excel
        with open(caminho, "w", newline="", encoding="utf-8-sig") as arq:
            escritor = csv.DictWriter(arq, fieldnames=cabecalhos)
            escritor.writeheader()
            escritor.writerows(linhas)
    except PermissionError as exc:
        raise PermissionError(
            f"Sem permissão para criar '{caminho}'. "
            "Verifique se a pasta está acessível e não está aberta por outro programa."
        ) from exc

    logger.debug("CSV exportado: %s", caminho)
    return caminho.resolve()


# ---------------------------------------------------------------------------
# Exportação para Excel
# ---------------------------------------------------------------------------

def exportar_excel(dados: list[dict], diretorio: str = DIRETORIO_SAIDA) -> Path:
    """
    Salva os resultados em um arquivo Excel (.xlsx) com formatação básica:
    cabeçalho em negrito, largura automática de colunas e filtros habilitados.

    Requer apenas openpyxl (`pip install openpyxl`). A planilha é montada
    diretamente pela biblioteca, sem pandas: a única coisa que o pandas fazia
    aqui era transportar as linhas até o openpyxl, que já era quem escrevia o
    arquivo e aplicava os estilos. Carregar pandas e numpy só para isso
    acrescenta cerca de 100 MB ao executável distribuído.

    Args:
        dados: Lista de dicts retornada por BuscadorProvedores.buscar_todos().
        diretorio: Pasta de destino.

    Returns:
        Caminho absoluto do arquivo criado.

    Raises:
        ImportError: Se openpyxl não estiver instalado.
        PermissionError: Se o arquivo estiver aberto no Excel ou a pasta bloqueada.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError(
            "Para exportar em Excel instale: pip install openpyxl"
        ) from exc

    pasta = Path(diretorio)
    _garantir_diretorio(pasta)

    caminho = pasta / f"{_nome_arquivo_base()}.xlsx"
    cabecalhos = list(COLUNAS_SAIDA.values())
    linhas = _preparar_linhas(dados)

    arquivo = Workbook()
    planilha = arquivo.active
    planilha.title = "Provedores"

    planilha.append(cabecalhos)
    for linha in linhas:
        # Números seguem como números, e não como texto: assim o Excel ordena
        # a coluna de distância por grandeza e permite filtrar por faixa.
        planilha.append([linha[rotulo] for rotulo in cabecalhos])

    # ---- Estilo do cabeçalho ----
    cor_cabecalho = "1F4E79"  # azul escuro
    for celula in planilha[1]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = PatternFill("solid", fgColor=cor_cabecalho)
        celula.alignment = Alignment(horizontal="center")

    # ---- Largura automática ----
    for indice, rotulo in enumerate(cabecalhos, start=1):
        maior = max(
            [len(str(rotulo))] + [len(str(linha[rotulo])) for linha in linhas]
        )
        # Limita a largura máxima a 60 caracteres
        planilha.column_dimensions[get_column_letter(indice)].width = min(maior + 2, 60)

    # ---- Filtros automáticos e cabeçalho congelado ----
    planilha.auto_filter.ref = planilha.dimensions
    planilha.freeze_panes = "A2"

    try:
        arquivo.save(caminho)
    except PermissionError as exc:
        raise PermissionError(
            f"Sem permissão para criar '{caminho}'. "
            "Verifique se o arquivo não está aberto no Excel."
        ) from exc

    logger.debug("Excel exportado: %s", caminho)
    return caminho.resolve()


# ---------------------------------------------------------------------------
# Função unificada
# ---------------------------------------------------------------------------

def exportar_resultados(
    dados: list[dict],
    formato: str = "csv",
    diretorio: str = DIRETORIO_SAIDA,
) -> list[Path]:
    """
    Exporta os resultados no(s) formato(s) solicitado(s).

    Args:
        dados: Lista de provedores retornada por BuscadorProvedores.buscar_todos().
        formato: "csv", "excel" ou "ambos".
        diretorio: Pasta de destino.

    Returns:
        Lista com os caminhos absolutos dos arquivos gerados.

    Raises:
        ValueError: Se o formato for inválido.
    """
    formatos_validos = {"csv", "excel", "ambos"}
    if formato not in formatos_validos:
        raise ValueError(
            f"Formato inválido: '{formato}'. "
            f"Use um dos seguintes: {', '.join(sorted(formatos_validos))}"
        )

    if not dados:
        logger.warning("Nenhum dado para exportar.")
        return []

    arquivos: list[Path] = []

    if formato in ("csv", "ambos"):
        arquivos.append(exportar_csv(dados, diretorio))

    if formato in ("excel", "ambos"):
        arquivos.append(exportar_excel(dados, diretorio))

    return arquivos
